import os 

from openpyxl import load_workbook
import pandas as pd

from collections import Counter


location = f'C:/Users/mrodziew/OneDrive - Capgemini/Documents/1. AWS/1. Reports/5. Data Analysis'
dst = f'C:/Users/mrodziew/OneDrive - Capgemini/Documents/1. AWS/1. Reports/5. Data Analysis'

def inc_search():
    """search string from the INC numbers into txt file"""
    try:
        wb = load_workbook(f'{location}/incs.xlsx')
        ws_inc = wb['INC']
        with open(f"{dst}/INC.txt", 'w') as inc_txt:
            for row in ws_inc.rows:
                for cell in row:
                    inc_txt.write(f"""'Incident ID*+' = "{cell.value}" OR """)
            inc_txt.seek(0, 2)
            inc_txt.seek(inc_txt.tell() - 3, 0)
            inc_txt.truncate()
    except FileNotFoundError:
        print('no such file')


def wo_search():
    """search string from the REQ numbers into txt file"""
    try:
        wb = load_workbook(f'{location}/wos.xlsx')
        ws_inc = wb['WO']
        with open(f"{dst}/WO.txt", 'w') as inc_txt:
            for row in ws_inc.rows:
                for cell in row:
                    inc_txt.write(f"""'Associated Request ID' = "{cell.value}" OR """)
            inc_txt.seek(0, 2)
            inc_txt.seek(inc_txt.tell() - 3, 0)
            inc_txt.truncate()
    except FileNotFoundError:
        print('no such file')


def inc_wo_req_search():
    """search from req_inc and req_wo numbers into txt files"""
    wb = load_workbook(f"{location}/data.xlsx")
    ws_inc = wb['INC']
    ws_wo = wb['WO']
    with open(f"{dst}/INC.txt", 'w') as inc_txt:
        for row in ws_inc.rows:
            for cell in row:
                inc_txt.write(f"""'Service Request ID' = "{cell.value}" OR """)

        inc_txt.seek(0, 2)
        inc_txt.seek(inc_txt.tell() - 3, 0)
        inc_txt.truncate()

    with open(f"{dst}/WO.txt", 'w') as wo_txt:
        for row in ws_wo.rows:
            for cell in row:
                wo_txt.write(f"""'Associated Request ID' = "{cell.value}" OR """)

        wo_txt.seek(0,2)
        wo_txt.seek(wo_txt.tell() -3,0)
        wo_txt.truncate()


def _string_splitter():
    """Read excel file to get a list of factors"""
    try:
        df = pd.read_excel(f"{location}/HS_data.xlsx",sheet_name='Sheet1')
        df_new = df[df['Factor'].notnull()]
        list_of_factors = df_new['Factor'].tolist()
        split_list_factors = []
        for item in list_of_factors:
            split_list_factors.extend(item.split(', '))
        return split_list_factors
    except FileNotFoundError:
        print('no such file')

        
def factor_counter(raw_list):
    """Get a list of factors and calculate statistics write to a file"""
    calculated_list = Counter(raw_list).most_common() 
    with open(f"{dst}/hs_calculations.txt", 'w') as hs_cal:
        for item in calculated_list:
            hs_cal.write(f'{item} \n')
    
